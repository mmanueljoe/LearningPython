import random

x = "Emmanuel Joe"
print(x[2])
print(x.count("m"))

car = {" name": "Corolla", " year": 2020, " color": " white ", " origin": "Ghana "}
print(car)

b = " Mathematics "
print(b[:5])
print(b.strip())
print(b)
print(b[2:])
print(random.randrange(20, 100))
c = 5j
print(type(c))
print(b.upper())
print(b.lower())
print(b.replace("h", "j"))
print(b)
v = "hello"
h = " world"
print(v + h)
age = 36
txt = "My name is Joe and I am {} years old"
print(txt.format(age))
t = "Ama is a 'beautiful' lady"
print(t)
n = "Ama is a \"bold\" wife"
print(n)
m = 'Ama is a "bold" lady'
print(m)
k = "Kojo is a winner\b"
print(k)
print(type(k))
text = "is it okay?"
print(text)
print(text.capitalize())
print("ok")
print(text.index("i"))
print(text.lower())
print(text.casefold())
print(text.find("is"))
print(text.find(" "))
print(text.count(" "))

rains = True
noRain = True
if rains:
    print("Go outside")
else:
    print("Stay Inside")

print(rains)
num = 45
print(str(num))

fruits_another = list(("apple", "cherry", "pawpaw", "cocoa"))
print(fruits_another)

fruits = ["apple", "pawpaw", "cherry", "cocoa", "melon", "tomato"]
print(fruits)
for fruit in fruits:
    print(fruit)
for x in range(9):
    print(x)
vegetables = ["okro", "cabbage", "carrot", "tomato"]
for x in fruits:
    if x in vegetables:
        print(x + " is a vegetable")
    else:
        print(x + " is a fruit")

for x in fruits:
    print(x[0])
for y in fruits:
    print(y[-1])
for x in fruits[0]:
    print(x)

print(len(fruits))
if "melon" in fruits:
    print("Melon is in the list of fruits")
else:
    print("Melon is not in the list of fruits")

fruits[1:3] = ["watermelon", "orange"]
print(fruits)
print(len(fruits))
fruits.insert(2, "mango")
print(fruits)
fruits.insert(7, "banana")
print(fruits)
print(fruits.index("banana"))
print(len(fruits))
fruits.extend(vegetables)
print(fruits)
print(vegetables)
print(fruits.count("tomato"))
print(fruits.pop())
print(fruits)
print(len(fruits))

x = 0
while x < len(fruits):
    print(fruits[x])
    x = x + 1
# list comprehension
newlist = []
for x in fruits:
    if x in vegetables:
        newlist.append(x)
        print(newlist)
x = 0
while x < len(fruits):
    print(fruits[0][0])
    x = x + 1

ran = ["toyota", "apple", "horse", "goat", "hen", "corolla", "pajaro", "mango", "avocado"]
Fruits = [x for x in ran if "a" in x]
print(Fruits)
print(ran)

Newlist = [x.capitalize() for x in ran]
print(Newlist)

New_list = [x for x in range(90)]
print(New_list)
values = [100, 500, 400, 300, 600, 200, 700, 800, 900, 1000, 10, 20, 30, 40, 50, 60, 70, 80, 90]
values.sort()
print(values)
print(values)

values.sort(reverse=True)
print(values)


def sayamen():
    print("Amen")


sayamen()


def tri_recursion(k):
    if k > 0:
        result = k + tri_recursion(k - 1)
        print(result)
    else:
        result = 0
    return result


print("\n\nRecursion Example Results")
tri_recursion(6)


class Theclass:
    name = "Joe"


# Assignment
"""name = input("Please Enter Your Name :")
age = input("Please Enter Your Age :")
Birthyear = 2022 - int(age)
def getinfo(name,age):
    print(format(f"Hello,{name}.You were born in the year {Birthyear}"))
getinfo(name,age)"""

"""NexGen = ["Godwin","Fredrick","Collins","Precious","Eli","Samuel","Prosper","Nabil","Seth","Ezekiel","Ernest","Sefakor","Ernest","Nk","Paa Kwesi","Prosper","Ignatus","Emmanuel","Israel","Edwin","Joe"]
wentForward = ["Ignatus","Edwin","Collins","Fredrick"]
def Printname():
    for x in NexGen[:]:
        if x in wentForward:
            NexGen.remove(x)
        else:
            print(x)
Printname()"""

NexGen = ["Godwin", "Fredrick", "Collins", "Precious", "Eli", "Samuel", "Prosper", "Nabil", "Seth", "Ezekiel", "Ernest",
          "Sefakor", "Ernest", "Nk", "Paa Kwesi", "Prosper", "Ignatus", "Emmanuel", "Israel", "Edwin", "Joe"]
wentForward = ["Ignatus", "Edwin", "Collins", "Fredrick"]
new_NexGen = []


def printName():
    for name in NexGen:
        if name not in wentForward:
            new_NexGen.append(name)
            print(name)


printName()

# Practising Python
# #guessing game
secret_number = 1
guessCount = 0
guesslimit = 3
"""while guessCount < guesslimit:
    guess = int(input("Guess"))
    guessCount += 1
    if guess == secret_number:
        print("You won")
        break
else:
    print("Sorry you failed!")"""

# car game
'''text = "Help"
enter = input("Need help?....Enter help: ")
if enter == text:
    print("Enter;\n"
          "1.Start - To Start the Car\n"
          "2.Stop to Stop the Car\n"
          "3.Quit to Exit")
enter1 = input(">")
if enter1.lower() == "Start":
    print("Car Started....Ready To Go!")
elif enter1.lower() == "Stop":
    print("Car Stopped...Bye!")
elif enter1.lower() == "Quit":
    print("Exiting")
else:
    print("I don't Understand that....")


command = ""
started = False
while True:
    command = input("> ").lower()
    if command == "start":
        if started:
          print("Car is already Started!")
        else:
            started = True
    elif command == "stop":
        if not started:
            print("Car is already stopped")
        else:
            started = False
            print("Car stopped")
        print("Car Stopped.")
    elif command == "help":
        print("""
        start - to start the car
        stop - to stop the car
        quit - to quit""")
    elif command == "quit":
        break
    else:
        print("Sorry ,I don't understand that... ")'''

# for loops
num = [1, 2, 3, 4, 5, 6]
for x in num:
    print(x, end="")
fruits_another = []
for y in range(0, 101):
    fruits_another.append(y)
    print(y)
print(fruits_another)

prices = [10, 20, 30]
add = 0
for z in prices:
    add += z
print(add)

# nested loops
# coordinates
for x in range(4):
    for y in range(3):
        print(f"({x},{y})")

numbers = [5, 2, 5, 2, 2]
for u in numbers:
    print("x" * u)

a_numbers = [5, 2, 5, 2, 2]
for t in a_numbers:
    output = ""
    for i in range(t):
        output += "x"
        print(output)

another_num = [2, 2, 2, 2, 8]
for x in another_num:
    print("x" * x)

'''values = [2,2,2,2,8]
for y in values:
    result = ""
    for h in range(y):
        result += 1
        print(result)'''

number_list = [1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18, 19, 20]
maximum = number_list[0]
for x in number_list:
    if x > maximum:
        maximum = x
print(maximum)

# 2d list
matrix = [
    [1, 2, 3],
    [4, 5, 6],
    [7, 8, 9]
]
print(matrix[0])
for j in matrix:
    for item in j:
        print(item)

# tuples
'''Tuples do not support changes.You can only get info about them'''
tuples = (1, 2, 3, 4)

# unpacking
coordinates = (1, 2, 3)
'''x = coordinates[0]
y = coordinates[1]
z = coordinates[2]'''  # simplified below
x, y, z = coordinates
print(y)

# dictionaries
dic = {"Name": "Joe", "Phone": "0546169405", "Gender": "Male", "Age": 20}
for a in dic:
    print(a)

# converting int to string
'''phone = input("Phone: ")
digits_mapping = {
    "0":"zero",
    "5":"five",
    "4":"four",
    "1": "one",
    "6":"six",
    "9": "nine"
}
output = " "
for num in phone:
    output += digits_mapping.get(num, " ! ") + " "
print(output)'''

# coverting emojis(dictionaries)
'''message= input("> ")
words = message.split(" ")
print(words)
emojis = {
    ":)":"smile",
}
print(type(words))'''

# functions

# handling errors(Exception)
try:
    '''age = int(input("Age: "))
    print(age)'''
except ZeroDivisionError:
    print("Age Cannot be Zero")
except ValueError:
    print("Invalid Value!")


# classes
class Point:
    def __init__(self, x, o):
        self.x = x
        self.y = o

    def move(self):
        print("move")

    def draw(self):
        print("draw")


point1 = Point(10, 30)
print(point1.x)
print(point1.y)
point1.draw()

point2 = Point(20, 30)
print(point2.x)
print(point2.y)


class Person:
    def __int__(self, name):
        self.name = name

    def talk(self):
        print("Talk")


joe = Person()
joe.name = "Joe"
print(joe.name)
joe.talk()


# Inheritance
class Mammals:
    def walk(self):
        print("walk")


class Dog(Mammals):
    pass


class Cat(Mammals):
    pass


german = Dog()
print(german.walk())

kitty = Cat()

# modules
from utils import find_max

values1 = [1, 2, 3, 4, 5, 6, 7, 8]
print(find_max(values1))

# packages
'''import ecommerce.shipping
ecommerce.shipping.shipping_fee()'''
# using the second approach
from ecommerce.shipping import shipping_fee

# random values
import random

'''for x in range(2):
    print(random.randint(0,2))'''

NexGen = ["Godwin", "Fredrick", "Collins", "Precious", "Eli", "Samuel", "Prosper", "Nabil", "Seth", "Ezekiel", "Ernest",
          "Sefakor", "Ernest", "Nk", "Paa Kwesi", "Prosper", "Ignatus", "Emmanuel", "Israel", "Edwin", "Joe", "sylvia"]
wentForward = ["Ignatus", "Edwin", "Collins", "Fredrick"]
leader = random.choice(NexGen)
print(leader)


# Dice
class Dice:
    def roll(self):
        first = random.randint(1, 6)
        second = random.randint(1, 6)
        return first, second


roll1 = Dice()
print(roll1.roll())


#trials
'''try:
    print("Enter A Character:")
    character = input()
    if character >= "A" and character <= "Z":
        print("Charcter is Upper case")
    elif character >= "a" and character <= "z":
        print("Character is Lower Case")
    else:
        print("Invalid Input!,Enter A Character")
except SyntaxError:
    print("Invalid Input")'''


#files and directories
from pathlib import Path
  #absolute path
# #relative path
path = Path("ecommerce")
print(path.exists())
'''path1 = Path("emails
print(path1.mkdir())''' # remove directory (rmdir)
# make a new directory(mkdir)'''
path2 = Path()
for file in path2.glob("*.py"):
    print(file)

#processing spreedsheets
'''import openpyxl as xl
from openpyxl.chart import BarChart,Reference
wb = xl.load_workbook("transactions.xlsx")
sheet = wb["Sheet1"]
cell = sheet["a1"]
cell = sheet.cell(1,1)
print(sheet.max_row())

for row in range(2,sheet.max_row +1):
    cell = sheet.cell(row,3)
    print(row)
    cprice = cell.value * 0.9
    cprice_cell = sheet.cell(row,4)
    cprice_cell.value = cprice

values = Reference(sheet,min_row=2,max_row=sheet.max_row,min_col=4,max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart,"e2")

wb.save = ('transactions2.xlsx')'''

#machine learning(AI)



